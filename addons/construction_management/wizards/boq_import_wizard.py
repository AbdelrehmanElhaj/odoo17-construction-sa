import base64
import io
from odoo import models, fields, api
from odoo.exceptions import UserError


class ConstructionBoqImport(models.TransientModel):
    _name = 'construction.boq.import'
    _description = 'BOQ Excel Import Wizard | استيراد BOQ من Excel'

    boq_id = fields.Many2one('construction.boq', string='BOQ', required=True)
    file_data = fields.Binary(string='ملف Excel أو CSV', required=True, attachment=False)
    file_name = fields.Char(string='اسم الملف')
    import_result = fields.Html(string='نتيجة الاستيراد', readonly=True)

    # ── Template download ─────────────────────────────────────────────

    def action_download_template(self):
        try:
            import openpyxl
            from openpyxl.styles import Font, PatternFill, Alignment
        except ImportError:
            raise UserError('مكتبة openpyxl غير مثبتة. استخدم قالب CSV.')

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = 'BOQ Items'

        headers = [
            'item_code', 'description', 'description_ar',
            'item_type', 'uom', 'qty_estimated', 'unit_price', 'notes',
        ]
        header_labels = [
            'كود البند', 'الوصف (EN)', 'الوصف (AR)',
            'النوع (material/labor/equipment/subcontract/overhead)',
            'وحدة القياس', 'الكمية التقديرية', 'سعر الوحدة', 'ملاحظات',
        ]
        header_fill = PatternFill('solid', fgColor='1F4E79')
        header_font = Font(color='FFFFFF', bold=True)

        ws.append(headers)
        ws.append(header_labels)

        for col_num, _ in enumerate(headers, start=1):
            cell = ws.cell(row=1, column=col_num)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center')

        # Sample rows
        sample_rows = [
            ['MAT-001', 'Portland Cement 50kg', 'أسمنت بورتلاند 50كغ', 'material', 'Bag', 500, 15.0, ''],
            ['MAT-002', 'Steel Rebar 12mm', 'حديد تسليح 12مم', 'material', 'Ton', 20, 2800.0, ''],
            ['LAB-001', 'Mason Daily Labor', 'عمال بناء', 'labor', 'Day', 30, 150.0, ''],
            ['EQP-001', 'Concrete Mixer', 'خلاطة خرسانة', 'equipment', 'Day', 10, 300.0, ''],
        ]
        for row in sample_rows:
            ws.append(row)

        ws.column_dimensions['B'].width = 30
        ws.column_dimensions['C'].width = 30
        ws.column_dimensions['D'].width = 40

        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)

        attachment = self.env['ir.attachment'].create({
            'name': 'boq_import_template.xlsx',
            'datas': base64.b64encode(buffer.read()),
            'mimetype': 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            'res_model': 'construction.boq.import',
            'res_id': self.id,
        })
        return {
            'type': 'ir.actions.act_url',
            'url': f'/web/content/{attachment.id}?download=true',
            'target': 'new',
        }

    # ── Import ────────────────────────────────────────────────────────

    def action_import(self):
        if not self.file_data:
            raise UserError('يرجى رفع ملف للاستيراد.')
        file_name = (self.file_name or '').lower()
        if file_name.endswith('.csv'):
            rows = self._read_csv()
        else:
            rows = self._read_excel()
        return self._process_rows(rows)

    def _read_excel(self):
        try:
            import openpyxl
        except ImportError:
            raise UserError('مكتبة openpyxl غير مثبتة. يرجى رفع ملف CSV.')
        data = base64.b64decode(self.file_data)
        wb = openpyxl.load_workbook(io.BytesIO(data), data_only=True)
        ws = wb.active
        all_rows = list(ws.iter_rows(values_only=True))
        if not all_rows:
            raise UserError('الملف فارغ.')
        # Skip up to 2 header rows (codes row + labels row)
        return all_rows[2:] if len(all_rows) > 2 else all_rows[1:]

    def _read_csv(self):
        import csv
        data = base64.b64decode(self.file_data).decode('utf-8-sig')
        reader = csv.DictReader(io.StringIO(data))
        return [
            (
                row.get('item_code', ''), row.get('description', ''),
                row.get('description_ar', ''), row.get('item_type', 'material'),
                row.get('uom', ''), row.get('qty_estimated', 0),
                row.get('unit_price', 0), row.get('notes', ''),
            )
            for row in reader
        ]

    def _process_rows(self, rows):
        VALID_TYPES = {'material', 'labor', 'equipment', 'subcontract', 'overhead'}
        UomModel = self.env['uom.uom']
        uom_cache = {}
        created = 0
        errors = []

        for i, row in enumerate(rows, start=3):
            if not row or not any(v for v in row if v is not None and str(v).strip()):
                continue

            def _cell(idx):
                try:
                    val = row[idx]
                    return str(val).strip() if val is not None else ''
                except IndexError:
                    return ''

            def _float(idx):
                try:
                    val = row[idx]
                    return float(val) if val not in (None, '') else 0.0
                except (ValueError, TypeError):
                    return 0.0

            description = _cell(1)
            if not description:
                errors.append(f'السطر {i}: الوصف (description) مطلوب — تم التخطي.')
                continue

            item_type = _cell(3).lower() or 'material'
            if item_type not in VALID_TYPES:
                item_type = 'material'

            uom_name = _cell(4)
            uom_id = False
            if uom_name:
                if uom_name not in uom_cache:
                    uom = UomModel.search([('name', 'ilike', uom_name)], limit=1)
                    uom_cache[uom_name] = uom.id if uom else False
                uom_id = uom_cache[uom_name]

            self.env['construction.boq.line'].create({
                'boq_id':       self.boq_id.id,
                'sequence':     (created + 1) * 10,
                'item_code':    _cell(0) or False,
                'description':  description,
                'description_ar': _cell(2) or False,
                'item_type':    item_type,
                'uom_id':       uom_id,
                'qty_estimated': _float(5),
                'unit_price':   _float(6),
                'notes':        _cell(7) or False,
            })
            created += 1

        result_lines = [f'<p class="text-success fw-bold">تم استيراد <strong>{created}</strong> بند بنجاح.</p>']
        if errors:
            result_lines.append('<p class="text-warning fw-bold">تحذيرات:</p><ul>')
            result_lines.extend(f'<li>{e}</li>' for e in errors)
            result_lines.append('</ul>')
        self.write({'import_result': ''.join(result_lines)})

        return {
            'type': 'ir.actions.act_window',
            'res_model': 'construction.boq.import',
            'res_id': self.id,
            'view_mode': 'form',
            'target': 'new',
        }
